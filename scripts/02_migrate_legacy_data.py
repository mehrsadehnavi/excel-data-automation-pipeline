# In this code, we want to extract data from the outdated files and put them in the new files.
# The data from the previous files is heterogeneous and needs to be modified.

import os
from openpyxl import load_workbook

# Paths to the folders
teacher_folder = "Project\\sample_data"     # Source
output_folder = "Project\\sample_data"     # Destination

# Row numbers from original files
original_rows = [22, 23, 24, 38, 39, 40, 41, 42, 59, 60, 61, 62, 63, 64, 65, 66, 79, 80, 81, 82, 83, 84, 85,
86, 87, 88, 89]

for filename in os.listdir(teacher_folder):
    if filename.endswith(".xlsx"):
        teacher_name = os.path.splitext(filename)[0]

        original_file = os.path.join(teacher_folder, filename)
        new_file = os.path.join(output_folder, f"{teacher_name}.xlsx")

        # Open both workbooks
        try:
            wb_old = load_workbook(original_file, data_only=True)
            wb_new = load_workbook(new_file)

            ws_old = wb_old["results"]
            ws_new = wb_new["Levels"]

            for old_row in original_rows:
                #C
                #value = ws_old[f"C{old_row}"].value
                #new_row = old_row
                #ws_new[f"C{new_row}"] = value
                #d
                value = ws_old[f"D{old_row}"].value
                new_row = old_row - 16
                ws_new[f"D{new_row}"] = value
                #E
                #value = ws_old[f"E{old_row}"].value
                #new_row = old_row
                #ws_new[f"E{new_row}"] = value
                #F
                #value = ws_old[f"F{old_row}"].value
                #new_row = old_row
                #ws_new[f"F{new_row}"] = value
                #G
                #value = ws_old[f"G{old_row}"].value
                #new_row = old_row
                #ws_new[f"G{new_row}"] = value
                #levels
                value = ws_old[f"E{old_row}"].value
                new_row = old_row
                ws_new[f"E{new_row}"] = value


            wb_new.save(new_file)
            print(f"Updated: {new_file}")

        except Exception as e: # In case one workbook was already opened and it couldn't procceed with the action.
            print(f"Error processing {teacher_name}: {e}")
