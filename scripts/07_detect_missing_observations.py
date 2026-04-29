import os
import openpyxl


folder_path = 'Project\\sample_data'
teacher_rows = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]  # Adjusted for openpyxl row indices

teachers_with_empty_observe = []

# Loop through all Excel files in the specified folder
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsx') and filename != 'temp.xlsx':
        teacher_name = filename.split('.')[0]
        teacher_file_path = os.path.join(folder_path, filename)

        try:
            # Load the workbook and select the "results" sheet
            workbook = openpyxl.load_workbook(teacher_file_path, data_only=True)
            if "results" not in workbook.sheetnames:
                print(f"'results' sheet not found in {teacher_file_path}")
                continue
            sheet = workbook["results"]

            # Extract the specified rows from column D (4th column, index 3)
            observe_values = [
                sheet.cell(row=row + 1, column=4).value for row in teacher_rows
            ]

            # Check if all values are empty or invalid
            non_numeric_or_nan = all(
                value is None or isinstance(value, str) and value.strip() in ["#DIV/0!", ""] or
                not isinstance(value, (int, float))
                for value in observe_values
            )

            if non_numeric_or_nan:  # If all values are invalid
                teachers_with_empty_observe.append(teacher_name)

        except Exception as e:
            print(f"Error processing {teacher_file_path}: {e}")

# Save the teacher names to a text file
output_file_path = os.path.join(folder_path, 'teachers_with_empty_observe.txt')
try:
    with open(output_file_path, 'w', encoding='utf-8') as file:
        file.write("\n".join(teachers_with_empty_observe))
    print(f"Teacher names successfully saved to: {output_file_path}")
except Exception as e:
    print(f"Error saving teacher names to file: {e}")

# Print the list of teachers with empty or invalid "Observe" column data
print("Teachers with empty or invalid 'Observe' column data:")
print(teachers_with_empty_observe)
