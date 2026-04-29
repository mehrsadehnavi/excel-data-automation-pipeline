import pandas as pd
import os
from openpyxl import load_workbook
import win32com.client

# Define folder paths and password
source_folder_path = 'Project\\sample_data'  
destination_folder_path = 'Project\\sample_data'  
temp_file_path = 'Project\\sample_data\\temp.xlsx'  

# Define rows to read and write
teacher_rows = [22, 23, 24, 38, 39, 40, 41, 42, 59, 60, 61, 62, 63, 64, 65, 66,
               79, 80, 81, 82, 83, 84, 85, 86, 87, 88]  # Rows to read in teacher files
result_rows = [22, 23, 24, 38, 39, 40, 41, 42, 59, 60, 61, 62, 63, 64, 65, 66,
               79, 80, 81, 82, 83, 84, 85, 86, 87, 88]  # Rows to write in temp.xlsx

# Ensure the destination folder exists
os.makedirs(destination_folder_path, exist_ok=True)


# Loop through all Excel files in the specified folder
for filename in os.listdir(source_folder_path):
    if filename.endswith('.xlsx') and filename != 'temp.xlsx':
        teacher_name = filename.split('.')[0]
        source_file_path = os.path.join(source_folder_path, filename)
        destination_file_path = os.path.join(destination_folder_path, filename)  # Destination path for the file

        # Check if the destination file exists; if not, create it
        if not os.path.exists(destination_file_path):
            # Create a blank file using the temp file structure
            temp_workbook = load_workbook(temp_file_path, read_only=False, keep_vba=False, data_only=False)
            temp_workbook.save(destination_file_path)
            temp_workbook.close()

        # Open the teacher file and collect data
        teacher_data = pd.read_excel(source_file_path, sheet_name=0, header=None, usecols="E",
                                     skiprows=[i for i in range(teacher_rows[0] - 1)]).iloc[
            [i - teacher_rows[0] for i in teacher_rows], 0].tolist()

        # Open the corresponding destination file
        destination_workbook = load_workbook(destination_file_path, read_only=False, keep_vba=False, data_only=False)

        # Write data into the "results" tab in the destination file
        results_sheet = destination_workbook['result']
        for i, row in enumerate(result_rows):
            results_sheet[f'E{row}'] = teacher_data[i] if i < len(teacher_data) else None  # Write data or leave blank

        # Save the updated destination file
        destination_workbook.save(destination_file_path)

        # Close the workbook
        destination_workbook.close()

        # Reapply the password to the destination file (optional)
        excel = win32com.client.Dispatch("Excel.Application")
        workbook = excel.Workbooks.Open(destination_file_path)
        workbook.SaveAs(destination_file_path)
        workbook.Close()
        excel.Quit()

print("Processing complete.")
